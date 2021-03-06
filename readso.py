#!c:\Python27\python.exe
import os
import slate
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ImageXL
from openpyxl.styles import Font
import csv
import datetime
import barcode
from barcode.writer import ImageWriter
from PIL import Image
import Tkinter,tkFileDialog
import ctypes  # An included library with Python install.
def Mbox(title, text, style):
    rc = ctypes.windll.user32.MessageBoxW(0, text, title, style)
    return rc
##  Styles:
##  0 : OK
##  1 : OK | Cancel
##  2 : Abort | Retry | Ignore
##  3 : Yes | No | Cancel
##  4 : Yes | No
##  5 : Retry | No 
##  6 : Cancel | Try Again | Continue

## IDABORT 3
## IDCANCEL 2
## IDCONTINUE 11
## IDIGNORE 5
## IDNO 7
## IDOK 1
## IDRETRY 4
## IDTRYAGAIN 10
## IDYES 6


root = Tkinter.Tk()
root.withdraw()
filename = tkFileDialog.askopenfilename(filetypes=[("PDF file","*.pdf")],title='Choose an SO PDF file')
if filename != None:
    responseno = Mbox(u'File selected', filename, 1)
    if ( responseno == 2 ):
        exit()

travelcust={}
travelcust["01 CTI (Cent)"]="01 CTI"
travelcust["02 Jabil Wolf"]="02 Wolfe"
travelcust["03 Assured MFG"]="03 Assured"
travelcust["04 TRU Machining"]="04 TRU"
travelcust["05 OTIS (Ewa)"]="05 OTIS"
travelcust["06 ThermoFisher (Ewa)"]="06 Thermo"
travelcust["07 Vecco (Luy Danh)"]="07 Vecco"
travelcust["08 Intevac (Luy Danh)"]="08 Intevac"
travelcust["09 Dung Marina"]="09 Marina"
travelcust["10 C.X.O (Sam)"]="10-C"
travelcust["11 Embraer (Sam)"]="11 Embraer"
travelcust["12 MW Technology - Cris (Sam)"]="12 MW Tech"

#filename = "C:\PT&R\\01 Customers\\02 Jabil Wolf\\01 Job Documents\SO 31878 PO 131400-E5212T\SO 31878.pdf"
txtfile = os.path.splitext(filename)[0]+'.txt'
out = open(txtfile,"w") 
with open(filename, 'rb') as f:
    doc = slate.PDF(f)
for page in doc[:2]:
    out.write(page)
out.close()
sodir = os.path.dirname(filename)
jobdir = os.path.dirname(sodir) # new
custdir = os.path.dirname(jobdir) # new
traveldir = sodir  # new
#custdir = os.path.dirname(sodir) # old, will be removed
#traveldir = custdir + "\\Traveler" # old, will be removed
#jobdir = custdir + "\\01 Job Documents"  # old, will be removed
faidir = sodir + "\\FAI"
shipdir = sodir + "\\Shipping"
custname = os.path.basename(custdir)
text_file = open(txtfile, "r")
lines = text_file.read().split('\n\n')
enum_lines = enumerate(lines)

found_item_no = 0
itemNameFound = 0
sosize=6
if (sosize > 2):
    adj1 = 0
    adj2 = 0
    lineSearch = 'Amount'
if (sosize == 2):
    adj1 = 7
    adj2 = 1
    lineSearch = 'Line'
if (sosize == 1):
    adj1 = 12
    adj2 = 1
    lineSearch = 'Item'


for iteration in enum_lines:
    index, item = iteration
    print (iteration)
    if  (item == "S.O. No."):
        sodate = lines[index+1]
        SO = lines[index+2]
        print sodate,SO
    if  (item == "P.O. No."):
        servicetype = lines[index+1]
        PO = lines[index+2]
        SOPO = "SO "+SO+" PO "+PO
        print servicetype, PO
    if  (item == "Delivery Date"):
        paymentDue = lines[index+1]
        soShipDate = lines[index+2]
        loShipDate = lines[index+4]
        print paymentDue, soShipDate, loShipDate
    if ( found_item_no > 0 ):
        if ( item.isdigit() ):
            itemNo = int(item)
            #print found_item_no,item
            if ( found_item_no == itemNo ):
                itemMax = itemNo
                found_item_no = itemNo + 1
            else:
                print itemMax
                itemNameFound = 1
                found_item_no = -1
        else:
            print itemMax
            itemNameFound = index
            found_item_no = -1
    if  (item == lineSearch ):
        #print item
        found_item_no = 1
    if ( itemNameFound ):
        index1 = index
        indexnext = index1 + itemMax
        itemNames = lines[index1:indexnext]
        print itemNames
        index1 = indexnext + adj1
        indexnext = index1 + itemMax
        itemDesc = lines[index1:indexnext]
        print itemDesc
        index1 = indexnext + adj2
        indexnext = index1 + itemMax
        itemQty = lines[index1:indexnext]
        print itemQty
        itemNameFound = 0
#
# Generate Directories Traveler
#
travelsodir = traveldir + "\\Traveler" #old
travelsodir = traveldir  #new
dirs = [faidir,shipdir,travelsodir]
for onedir in dirs:
    if not os.path.exists(onedir):
        os.makedirs(onedir)

print traveldir
for index,partNo in enumerate(itemNames):
    print index,partNo
    lineno = index + 1
    wb = load_workbook("c:\data\Traveler Template_Updated20170316.xlsx")
    ws = wb.active

    itemDescLines = itemDesc[index].split('\n')
    n = len(itemDescLines)
    if (n == 1):
        print "Description ",itemDescLines," not in correct format"
        exit()
    for i in range(n):
        if (i>0):
            desc = itemDescLines[i]
            if re.match(r'^Mat: ', desc):
                mtype = desc.replace("Mat: ", "", 1)
            if re.match(r'^Dwg: ', desc):
                drawno = desc.replace("Dwg: ", "", 1)
            if re.match(r'^Finish: ', desc):
                fin = desc.replace("Finish: ", "", 1)
        else:       
            pdescription = itemDescLines[0]
    
    shipdate=soShipDate
    duedate=paymentDue
    quoteno=""
    #PO="PO 123"
    jobno=SO+"-LN"+str(lineno)
    #drawno = "drawno"
    custno = travelcust[custname]
    qty_order = itemQty[index]
    #mtype = "M type"
    mlotno = "Lot 30304"
    outside = ""
    o_vendor = ""
    o_vendor_po = ""

    ws["C1"] = partNo 
    ws["C2"] = pdescription
    d = datetime.datetime.now()
    only_date, only_time = d.date(), d.time()
    ws['I3'].number_format = 'm/d/y'
    ws["I3"] = shipdate
    ws["I4"] = duedate
    ws['I2'] = "=DATE(YEAR(I3), MONTH(I3), DAY(I3)-3)"
    #ws["I3"] = quoteno
    ws["I6"] = PO
    ws["H9"] = fin
    ws["H10"] = o_vendor
    ws["H11"] = o_vendor_po
    ws["C5"] = jobno
    ws["C6"] = custno
    ws["C7"] = drawno
    ws["C8"] = only_date
    ws['C8'].number_format = 'm/d/y'
    ws["C10"] = qty_order
    ws["C14"] = mtype
    c24 = ws['C24']
    c25 = ws['C25']
    c26 = ws['C26']
    c27 = ws['C27']
    c28 = ws['C28']
    ft = Font(size=11)
    c24.font = ft
    c25.font = ft
    c26.font = ft
    c27.font = ft
    c28.font = ft
    #ws["C14"] = mlotno
    imgfile = travelsodir + "\\" + partNo
    pngfile = travelsodir + "\\" + partNo + ".png"
    CODE39 = barcode.get_barcode_class('code39')
    code39 = CODE39(partNo, writer=ImageWriter(),add_checksum=False)
    fullname = code39.save(imgfile)
    img = Image.open(fullname) # image extension *.png,*.jpg
    new_width  = 186
    new_height = 120
    img = img.resize((new_width, new_height), Image.ANTIALIAS)
    #img.save(fullname)
    img0 = ImageXL(fullname)
    ws.add_image(img0,'A29')
    img1 = ImageXL('c:\\data\\logo1.png')
    ws.add_image(img1,'L1')
    img2 = ImageXL('c:\\data\\logo2.png')
    ws.add_image(img2,'L4')
    traveler = travelsodir + "\\T " + jobno + ".xlsx"
    wb.save(traveler)
    wb.close() 
Mbox(filename, u'Travelers were succesfully created', 0)	
